#!/usr/bin/env python
# -*- coding: utf-8 -*-
import argparse
import sys
import string
import signal
import os
import re
import json
from openpyxl import load_workbook
from enum import Enum
import time
import configparser


def siginit(sigNum, sigHandler):
    print("byebye")
    sys.exit(1)


signal.signal(signal.SIGINT, siginit)  # Ctrl-c处理


def openFile(filepath, mode, encoding):
    if sys.version_info.major == 3:
        return open(filepath, mode, encoding=encoding)
    else:
        import codecs
        return codecs.open(filepath, mode, encoding)


def log(msg):
    local_time = time.localtime(time.time())
    time_stamp = time.strftime('[%Y-%m-%d %H:%M:%S]', local_time)
    print(time_stamp + ' ' + msg)


def padding(_str, space_count):
    return _str.rjust(len(_str) + space_count)

# 主从表关系


class SheetType(Enum):
    # 普通表
    NORMAL = 0
    # 有主外键关系的主表
    MASTER = 1
    # 有主外键关系的附表
    SLAVE = 2


# 当有#id类型的时候  表输出json的是map形式(id:{xx:1})
# 当没有#id类型（或有#id[]）的时候  表输出json的是数组类型
class IndexType(Enum):
    ARRAY = 1
    MAP = 2


# 支持的数据类型
class DataType():
    NUMBER = 'number'
    STRING = 'string'
    BOOL = 'bool'
    DATE = 'date'
    ID = 'id'
    IDS = 'id[]'
    ARRAY = '[]'
    OBJECT = '{}'
    UNKNOWN = 'unknown'


DEFAULT_CONFIG = {
    'head_row': 2
}


class XlsxConfigHandler(object):
    def __init__(self, config, infile, outfile, export_type='all'):
        self.config = config                        # 配置
        self.infile = os.path.abspath(infile)       # excel文件名
        self.outfname = None
        self.split_output = False
        if outfile == None:
            f_dir, f_name = os.path.split(infile)
            self.outdir = f_dir
            if f_name != None:
                self.outfname = os.path.splitext(f_name)[0]
        else:
            f_dir, f_name = os.path.split(outfile)  # 导出data文件名
            self.outdir = f_dir
            if f_name != None:
                self.outfname = os.path.splitext(f_name)[0]
        self.export_type = export_type              # 导出data类型(json,js,ts)
        self.define_settings = {}
        self.parse()

    def parse(self):
        wb = load_workbook(self.infile)
        self.define_settings = self.parse_define(wb)
        self.sheet_datas = self.parse_sheet(wb, self.define_settings)
        wb.close()

    # 不需要解析的字段
    def is_field_no_parsed(self, name):
        return name.startswith('!') or name.startswith('！')

    def parse_define(self, workbook):
        log('=============start parse sheet define=============')
        head_row = self.config['head_row']
        settings = {}
        for sheet in workbook:
            log('parse sheet define:' + sheet.title)
            sheet_name = sheet.title
            if self.is_field_no_parsed(sheet_name):
                continue
            sheet_setting = {
                'name': sheet_name,
                'type': IndexType.ARRAY,
                'sheet_name': sheet_name,
                'sheet_type': SheetType.NORMAL,
                'master': None,
                'slaves': [],
                'head': []
            }
            is_slave = True if '@' in sheet_name else False
            if is_slave:
                sheet_setting['sheet_type'] = SheetType.SLAVE
                tmp_pair = sheet_name.split('@')
                sheet_name = tmp_pair[0]
                master_name = tmp_pair[1]
                sheet_setting['master'] = master_name
                settings[master_name]['slaves'].append(sheet_name)
                settings[master_name]['sheet_type'] = SheetType.MASTER
            row_data = sheet[head_row]
            for cell in row_data:
                if cell.value:
                    head_setting = {
                        'key': cell.value,
                        'type': DataType.UNKNOWN,
                        'number_format': cell.number_format,
                        'color': cell.fill.fgColor.rgb if cell.fill.fgColor.type == 'rgb' else None
                    }
                    if '#' in cell.value:
                        tmp_pair = cell.value.split('#')
                        head_setting['key'] = tmp_pair[0]
                        head_setting['type'] = tmp_pair[1]
                        if tmp_pair[1] == DataType.ID:
                            sheet_setting['type'] = IndexType.MAP
                        else:
                            sheet_setting['type'] = IndexType.ARRAY
                    sheet_setting['head'].append(head_setting)
            settings[sheet_name] = sheet_setting
        print(settings)
        log('==============parse sheet define end==============')
        return settings

    def get_major_item(self, head_settings):
        for item in head_settings:
            if item['type'] == DataType.ID or item['type'] == DataType.IDS:
                return item
        return None

    def get_major_key(self, head_settings):
        for item in head_settings:
            if item['type'] == DataType.ID or item['type'] == DataType.IDS:
                return item['key']
        return None

    def get_specify_item(self, sheet_data, key, value):
        if isinstance(sheet_data, dict):
            if value in sheet_data:
                return sheet_data[value]
        elif isinstance(sheet_data, list):
            for idata in sheet_data:
                if isinstance(idata, dict) and key in idata and value == idata[key]:
                    return idata

    def to_boolan(self, value):
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            value = value.lower()
            if value == 'true':
                return True
            elif value == 'false':
                return False
        return None

    def to_number(self, value):
        if isinstance(value, int) or isinstance(value, float):
            return value
        if isinstance(value, str):
            if re.match(r"^[\+-]?\d+$", value) != None:
                return int(value)
            elif re.match(r"^0x[0-9a-fA-F]+$", value) != None:
                return int(value, 16)
            elif re.match(r"^[\+-]?\d+\.\d+$", value) != None:
                return float(value)
            # elif re.match(r"^[\+-]?\d+(\.\d+)?[eE][\+-]?\d+$", value) != None:
            #     # 指数字符串暂不支持
            #     pass
            return None

    def to_string(self, value):
        if isinstance(value, str):
            if value.startswith('"') and value.endswith('"'):
                return value[1:-1]
        return str(value)

    def to_auto_type(self, value):
        # number boolean string
        ret_value = self.to_number(value)
        if ret_value == None:
            ret_value = self.to_boolan(value)
            if ret_value == None:
                ret_value = self.to_string(value)
        return ret_value

    def to_object(self, value):
        if isinstance(value, dict):
            return value
        if isinstance(value, str):
            if not value.startswith('{'):
                value = '{' + value
            if not value.endswith('}'):
                value = value + '}'
            try:
                return json.loads(value)
            except (TypeError, ValueError) as e:
                print(e)
        return None

    def to_array(self, value):
        if isinstance(value, list):
            return value
        else:
            value = str(value)
            if not value.startswith('['):
                value = '[' + value
            if not value.endswith(']'):
                value = value + ']'
            try:
                return json.loads(value)
            except (TypeError, ValueError) as e:
                print(e)
        return None

    def get_value(self, ori_value, to_type):
        ret_value = None
        if to_type == DataType.NUMBER:
            ret_value = self.to_number(ori_value)
        elif to_type == DataType.STRING:
            ret_value = self.to_string(ori_value)
        elif to_type == DataType.BOOL:
            ret_value = self.to_boolan(ori_value)
        elif to_type == DataType.ID or to_type == DataType.IDS:
            ret_value = self.to_string(ori_value)
        elif to_type == DataType.OBJECT:
            ret_value = self.to_object(ori_value)
        elif to_type == DataType.ARRAY:
            ret_value = self.to_array(ori_value)
        else:
            ret_value = self.to_auto_type(ori_value)
        return ret_value

    def parse_row(self, row_data, head_settings):
        result = {}
        count = min(len(row_data), len(head_settings))
        for i in range(0, count):
            cell = row_data[i]
            define = head_settings[i]
            key = define['key']
            if self.is_field_no_parsed(key):
                continue
            result[key] = self.get_value(cell.value, define['type'])
        return result

    def parse_sheet(self, workbook, settings):
        log('=============start parse sheet content=============')
        head_row = self.config['head_row']
        sheet_datas = {}
        for sheet in workbook:
            sheet_name = sheet.title
            log('parse sheet content:' + sheet_name)
            if self.is_field_no_parsed(sheet_name):
                continue
            is_slave = True if '@' in sheet_name else False
            if is_slave:
                sheet_name = sheet_name.split('@')[0]
            sheet_setting = settings[sheet_name]
            datas = None
            if sheet_setting['type'] == IndexType.MAP:
                datas = {}
                k_key = self.get_major_key(sheet_setting['head'])
                if k_key:
                    for row in range(head_row+1, sheet.max_row+1):
                        data = self.parse_row(
                            sheet[row], sheet_setting['head'])
                        k_value = data[k_key]
                        datas[k_value] = data
                else:
                    print('no major key')
            else:
                datas = []
                for row in range(head_row+1, sheet.max_row+1):
                    datas.append(self.parse_row(
                        sheet[row], sheet_setting['head']))

            sheet_datas[sheet_name] = datas
        print(sheet_datas)

        for name in settings:
            log('check sheet type:%12s - %s' %
                (name, settings[name]['sheet_type']))
            if settings[name]['sheet_type'] == SheetType.MASTER:
                master_sheet = sheet_datas[name]
                master_setting = settings[name]
                master_major_key = self.get_major_key(master_setting['head'])
                slaves = settings[name]['slaves']
                log('find slave sheet:%12s - %s' %
                    (name, slaves))
                for sl_name in slaves:
                    slave_sheet = sheet_datas[sl_name]
                    slave_setting = settings[sl_name]
                    major_item = self.get_major_item(slave_setting['head'])
                    #   slave 表中所有数据
                    if major_item:
                        k_key = major_item['key']
                        k_type = major_item['type']
                        is_dict = isinstance(slave_sheet, dict)
                        for k in slave_sheet:
                            if is_dict:
                                data = slave_sheet[k]
                            else:
                                data = k
                            k_value = data[k_key]
                            del data[k_key]
                            master_item = self.get_specify_item(
                                master_sheet, master_major_key, k_value)
                            if master_item != None:
                                if k_type == DataType.IDS:
                                    # array
                                    if sl_name not in master_item:
                                        master_item[sl_name] = []
                                    master_item[sl_name].append(data)
                                else:
                                    # dict
                                    master_item[sl_name] = data
                        del sheet_datas[sl_name]
        print(sheet_datas)
        log('==============parse sheet content end==============')
        return sheet_datas

    def to_ts_type(self, data_type):
        TS_TYPE_DICT = {}
        TS_TYPE_DICT[DataType.NUMBER] = 'number'
        TS_TYPE_DICT[DataType.STRING] = 'string'
        TS_TYPE_DICT[DataType.BOOL] = 'boolean'
        TS_TYPE_DICT[DataType.DATE] = 'string'
        TS_TYPE_DICT[DataType.ID] = 'string'
        TS_TYPE_DICT[DataType.IDS] = 'string'
        TS_TYPE_DICT[DataType.ARRAY] = 'any[]'
        TS_TYPE_DICT[DataType.OBJECT] = 'any'
        TS_TYPE_DICT[DataType.UNKNOWN] = 'any'
        return TS_TYPE_DICT[data_type]

    def get_dts_str(self, name, setting):
        declare_str = 'declare interface ' + name + ' {\n'
        head_list = setting['head']
        for info in head_list:
            sheet_name = info['key']
            sheet_type = info['type']
            if sheet_name.startswith('!'):
                continue
            type_des = self.to_ts_type(sheet_type)
            declare_str += '    ' + sheet_name + ': ' + type_des + ';\n'
        slaves = setting['slaves']
        for slave_name in slaves:
            declare_str += '    ' + slave_name + ': ' + 'any' + ';\n'
        declare_str += '}\n'
        return declare_str

    def gen_dts_file(self):
        if not os.path.exists(self.outdir):
            os.makedirs(self.outdir)
        if self.split_output:
            for k in self.define_settings:
                setting = self.define_settings[k]
                if setting['sheet_type'] != SheetType.SLAVE:
                    data = self.get_dts_str(k, setting)
                    fp = openFile(os.path.join(
                        self.outdir, k + '.d.ts'), 'a+', 'utf-8')
                    fp.seek(0)
                    fp.truncate()
                    fp.write(data)
                    fp.flush()
                    fp.close()
        else:
            data = ''
            for k in self.define_settings:
                setting = self.define_settings[k]
                if setting['sheet_type'] != SheetType.SLAVE:
                    data += self.get_dts_str(k, setting) + '\n\n'

            fp = openFile(os.path.join(
                self.outdir, self.outfname + '.d.ts'), 'a+', 'utf-8')
            fp.seek(0)
            fp.truncate()
            fp.write(data)
            fp.flush()
            fp.close()

    def gen_json_file(self):
        try:
            if not os.path.exists(self.outdir):
                os.makedirs(self.outdir)
            if self.split_output:
                for k in self.sheet_datas:
                    data = json.dumps(
                        self.sheet_datas[k], indent=2, ensure_ascii=False)
                    fp = openFile(os.path.join(
                        self.outdir, k + '.json'), 'a+', 'utf-8')
                    fp.seek(0)
                    fp.truncate()
                    fp.write(data)
                    fp.flush()
                    fp.close()
            else:
                data = json.dumps(self.sheet_datas, indent=2,
                                  ensure_ascii=False)
                fp = openFile(os.path.join(
                    self.outdir, self.outfname + '.json'), 'a+', 'utf-8')
                fp.seek(0)
                fp.truncate()
                fp.write(data)
                fp.flush()
                fp.close()

        except (TypeError, ValueError) as e:
            print(e)

    def get_data_ts_str(self, data, key='', space_count=0):
        _str = ''
        if data == None:
            if len(key) > 0:
                return padding(key + ': null', space_count)
            else:
                return padding('null', space_count)
        if isinstance(data, dict):
            if len(key) > 0:
                _str = key + ': {\n'
            else:
                _str = '{\n'
            _str = padding(_str, space_count)
            for k in data:
                _str += self.get_data_ts_str(data[k], k, space_count+4) + ',\n'
            _str += padding('}', space_count)
        elif isinstance(data, list):
            if len(key) > 0:
                _str = key + ': [\n'
            else:
                _str = '[\n'
            _str = padding(_str, space_count)
            for d in data:
                _str += self.get_data_ts_str(d, '', space_count+4) + ',\n'
            _str += padding(']', space_count)
        elif isinstance(data, str):
            if len(key) > 0:
                _str = ('%s: "%s"') % (key, data)
            else:
                _str = '"' + data + '"'
            _str = padding(_str, space_count)
        elif isinstance(data, int) or isinstance(data, float):
            if len(key) > 0:
                _str = ('%s: %s') % (key, data)
            else:
                _str = str(data)
            _str = padding(_str, space_count)
        elif isinstance(data, bool):
            b_str = 'true' if data else 'false'
            if len(key) > 0:
                _str = ('%s: %s') % (key, b_str)
            else:
                _str = b_str
            _str = padding(_str, space_count)

        return _str

    def gen_ts_file(self):
        if not os.path.exists(self.outdir):
            os.makedirs(self.outdir)
        if self.split_output:
            for k in self.sheet_datas:
                head_str = 'export let ' + k + ' = '
                if isinstance(self.sheet_datas[k], dict):
                    head_str = 'export let ' + k + ': any = '
                elif isinstance(self.sheet_datas[k], list):
                    head_str = 'export let ' + k + ': any[] = '
                data = self.get_data_ts_str(self.sheet_datas[k])
                data = head_str + data
                print(data+'\n')
                fp = openFile(os.path.join(
                    self.outdir, k + '.ts'), 'a+', 'utf-8')
                fp.seek(0)
                fp.truncate()
                fp.write(data)
                fp.flush()
                fp.close()
        else:
            moudle_name = self.outfname.replace('_', '')
            data = 'export module ' + moudle_name + ' {\n'
            for k in self.sheet_datas:
                if k == 'demo':
                    print(self.sheet_datas[k])
                head_str = '    export function ' + k + '() {\n'
                if isinstance(self.sheet_datas[k], dict):
                    head_str = '    export function ' + k + '(): any {\n'
                elif isinstance(self.sheet_datas[k], list):
                    head_str = '    export function ' + k + '(): any[] {\n'
                body_str = self.get_data_ts_str(self.sheet_datas[k], '$', 8)
                body_str = body_str.replace('$:', 'return')
                data += head_str + body_str + '\n'
                data += '    }\n\n'
            data += '}'
            fp = openFile(os.path.join(
                self.outdir, self.outfname + '.ts'), 'a+', 'utf-8')
            fp.seek(0)
            fp.truncate()
            fp.write(data)
            fp.flush()
            fp.close()

    def export(self, type='all'):
        self.gen_json_file()
        self.gen_dts_file()
        # self.gen_ts_file()


def list_all_files(rootdir):
    files = []
    _list = os.listdir(rootdir)
    for i in range(0, len(_list)):
        path = os.path.join(rootdir, _list[i])
        if os.path.isdir(path):
            files.extend(list_all_files(path))
        if os.path.isfile(path):
            files.append(path)
    return files


def exportConfig(args):
    inpath = args.inpath
    outpath = args.outpath

    cfg_parser = configparser.ConfigParser()
    cfg_parser.read("config.ini", encoding="utf-8")
    cfg_parser.options('define')
    config = {}
    config['head_row'] = cfg_parser.getint('define', 'head_row')
    config['head_row'] = DEFAULT_CONFIG['head_row']
    if inpath == None:
        cfg_parser.options('file')
        inpath = cfg_parser.get('file', 'in_dir')
        outpath = cfg_parser.get('file', 'out_dir')
    if os.path.isdir(inpath):
        files = list_all_files(inpath)
        for i in files:
            out = i.replace(inpath, outpath)
            handler = XlsxConfigHandler(config, i, out)
            handler.export()
    elif os.path.isfile(inpath) and os.path.isfile(outpath):
        handler = XlsxConfigHandler(config, inpath, outpath)
        handler.export()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('-i', '--inpath', help='input path(file or dir)',
                        required=False, default=None)
    parser.add_argument('-o', '--outpath', help='out path(file or dir)',
                        required=False, default=None)
    args = parser.parse_args()
    print(args)
    exportConfig(args)


if __name__ == '__main__':
    main()
